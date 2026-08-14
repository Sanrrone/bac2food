"""Walk every sheet of an Excel workbook, classify each as spine / nutrient /
metadata / unknown. Surfaces sheets a v1 parser may have silently skipped.

Print the classification table at startup; any sheet flagged 'unknown' or
'metadata' that the user expected to contain nutrient data is an obvious
re-parse bug to fix.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Optional

import openpyxl
import pandas as pd

from . import header_detect


_FOOD_ID_NEEDLES = {n.lower() for n in header_detect.DEFAULT_NEEDLES}
_FOOD_NAME_NEEDLES = {"food name", "food", "name", "description", "ingredient",
                      "item name", "food item", "designation", "denomination"}
_METADATA_HINTS = {"list of tables", "table of contents", "notes", "readme",
                   "references", "abbreviations", "definitions", "legend",
                   "introduction", "factors"}


@dataclass
class SheetInfo:
    name: str
    n_rows: int
    n_cols: int
    header_row: Optional[int]
    has_food_id: bool
    has_food_name: bool
    n_numeric_cols: int
    classification: str   # "foods_spine" | "nutrient_sheet" | "metadata" | "unknown"
    sample_cols: list[str] = field(default_factory=list)
    note: str = ""


def _column_is_numeric(series: pd.Series) -> bool:
    """A column is 'numeric' if ≥ 60% of non-empty cells parse as a float."""
    vals = series.dropna()
    if len(vals) == 0:
        return False
    n_ok = 0
    for v in vals[:200]:
        try:
            float(str(v).replace(",", ".").replace("<", "").strip())
            n_ok += 1
        except (ValueError, TypeError):
            continue
    return n_ok / max(1, min(200, len(vals))) >= 0.6


def classify_workbook(xlsx_path: str | Path,
                      min_spine_rows: int = 50) -> list[SheetInfo]:
    """Return one SheetInfo per sheet."""
    p = Path(xlsx_path)
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    out: list[SheetInfo] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        n_rows, n_cols = ws.max_row or 0, ws.max_column or 0
        if n_rows <= 1 or n_cols <= 1:
            out.append(SheetInfo(sn, n_rows, n_cols, None, False, False, 0,
                                 "metadata", [], "tiny sheet"))
            continue
        header_row = header_detect.find_header_row(str(p), sheet=sn)
        try:
            df = pd.read_excel(p, sheet_name=sn, header=header_row if header_row is not None else 0,
                               nrows=300, engine="openpyxl")
        except Exception as e:
            out.append(SheetInfo(sn, n_rows, n_cols, header_row, False, False, 0,
                                 "unknown", [], f"read failed: {e}"))
            continue
        cols = [str(c) for c in df.columns]
        cols_lc = [c.strip().lower() for c in cols]
        has_food_id = any(c in _FOOD_ID_NEEDLES for c in cols_lc)
        has_food_name = any(c in _FOOD_NAME_NEEDLES for c in cols_lc)
        n_numeric = sum(_column_is_numeric(df[c]) for c in cols)
        sn_lc = sn.lower()
        is_meta_hint = any(h in sn_lc for h in _METADATA_HINTS)
        if is_meta_hint and not (has_food_id and n_numeric >= 5):
            cls = "metadata"
        elif has_food_id and has_food_name and n_rows >= min_spine_rows:
            cls = "foods_spine"
        elif has_food_id and n_numeric >= 3:
            cls = "nutrient_sheet"
        elif n_numeric >= 3 and n_rows >= min_spine_rows:
            cls = "nutrient_sheet"   # has data but no obvious id col — flag in note
        else:
            cls = "unknown"
        out.append(SheetInfo(sn, n_rows, n_cols, header_row,
                             has_food_id, has_food_name, n_numeric, cls,
                             cols[:8], ""))
    return out


def print_table(infos: list[SheetInfo], title: str = "") -> None:
    if title:
        print(f"\n[sheet-classify] {title}")
    print(f"  {'sheet':40s}  {'rows':>6}  {'cols':>5}  {'hdr':>4}  fid  fnm  numc  class")
    for s in infos:
        print(f"  {s.name[:40]:40s}  {s.n_rows:6d}  {s.n_cols:5d}  "
              f"{(str(s.header_row) if s.header_row is not None else '?'):>4s}  "
              f"{'Y' if s.has_food_id else '·':>3s}  "
              f"{'Y' if s.has_food_name else '·':>3s}  "
              f"{s.n_numeric_cols:>4d}  {s.classification}"
              + (f"  ({s.note})" if s.note else ""))
